"""
DeepMindQ — 45 Core Engine Enhancements Roadmap
===============================================
Generates an XLSX with all enhancements organized by engine.
"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    font_title, font_body, font_header, font_subheader, font_caption,
    align_header, align_text, align_number, align_date,
    fill_header, fill_data_row, border_header,
    NEUTRAL_200, NEUTRAL_900, NEUTRAL_600, PRIMARY, PRIMARY_LIGHT,
    COLUMN_WIDTHS, ROW_HEIGHTS, ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
)
from templates.palettes import get_palette

# ── Use a tech/futuristic palette to match the app's gold theme ──
use_palette_explicit("professional")

# ── Data ──
enhancements = [
    # ═══════════════════════════════════ EMAIL ENGINE ═══════════════════════════════════
    ("E-01", "Email Engine", "SMTP Provider Integration", "Integrate a real email sending provider (Resend, SendGrid, Amazon SES, or Postmark) with API key configuration in settings. The queue has status tracking but zero sending logic.", "Critical", "High", "Medium", "Sending"),
    ("E-02", "Email Engine", "Queue PATCH Endpoint", "The queue screen calls PATCH /api/queue with {action: 'pause'|'resume'} but only a GET handler exists. Add PATCH for pause/resume/retry/cancel actions.", "Critical", "High", "Low", "Queue"),
    ("E-03", "Email Engine", "Email Sending Cron Worker", "Implement a background cron or webhook-based worker that processes the SendQueue: picks pending items, calls the SMTP provider, updates status to sent/failed, handles retries with exponential backoff.", "Critical", "High", "Medium", "Sending"),
    ("E-04", "Email Engine", "Retry Logic with Exponential Backoff", "SendQueue has retryCount field but it's never incremented. On failure, retry up to 3 times with delays of 5min, 30min, 2hr. Mark as permanently failed after max retries.", "High", "Medium", "Low", "Queue"),
    ("E-05", "Email Engine", "Email Scheduling System", "Add date/time picker UI for scheduling emails. Store scheduledAt properly. The cron worker should respect scheduling and only send when scheduledAt <= now.", "High", "Medium", "Medium", "Queue"),
    ("E-06", "Email Engine", "Email Thread Tracking", "Add In-Reply-To and References headers for conversation threading. Track Message-ID in the Draft model. When a reply comes in, link it to the original draft.", "Medium", "Medium", "Medium", "Reply Mgmt"),
    ("E-07", "Email Engine", "Inbound Email Webhook (Replies)", "Create a webhook endpoint (POST /api/webhooks/reply) that receives parsed reply emails (via SendGrid inbound, Resend inbound, or custom MX). Auto-categorize replies (positive, negative, OOO, unsubscribe).", "Critical", "High", "Medium", "Reply Mgmt"),
    ("E-08", "Email Engine", "Bounce Webhook Handler", "Create POST /api/webhooks/bounce to receive bounce notifications from the email provider. Auto-classify as hard/soft bounce, update contact status, add to suppression list for hard bounces.", "High", "High", "Low", "Deliverability"),
    ("E-09", "Email Engine", "Unsubscribe Link & Compliance", "Auto-append an unsubscribe link to every outbound email. Create a GET /api/unsubscribe endpoint. Track consent status. Add List-Unsubscribe header. GDPR/CAN-SPAM compliant.", "Critical", "High", "Low", "Compliance"),
    ("E-10", "Email Engine", "A/B Testing for Subject Lines", "Generate 2-3 subject line variants per email using AI. Randomly assign to leads. Track open rates per variant. Declare a winner after N sends.", "Medium", "High", "High", "Optimization"),
    ("E-11", "Email Engine", "Email Open & Click Tracking", "Embed a 1x1 tracking pixel for opens. Rewrite links with tracking redirects for clicks. Store events in a new EmailEvent model. Show open/click rates in analytics.", "Medium", "High", "Medium", "Analytics"),
    ("E-12", "Email Engine", "Bulk Draft Operations", "Add bulk approve, reject, regenerate, and delete actions to the drafts screen. Support selecting multiple drafts via checkboxes and applying actions in batch.", "High", "Medium", "Low", "UX"),
    ("E-13", "Email Engine", "Deduplicate Generation Code", "Three copies of email generation logic exist: email-generation.ts, drafts/route.ts, ai/generate/route.ts. Consolidate into a single source of truth. Remove hardcoded ZAI_CONFIG credentials.", "High", "Medium", "Medium", "Code Quality"),
    ("E-14", "Email Engine", "Email Template System", "Create a template library with editable subject/body patterns per service line. Support variable placeholders like {{name}}, {{company}}, {{pain_point}}. Let users create custom templates.", "High", "Medium", "Medium", "Generation"),
    ("E-15", "Email Engine", "Multi-Step Email Sequences", "Design drip campaign support: Sequence = ordered list of emails with delays (e.g., Day 0: intro, Day 3: case study, Day 7: CTA). Track which step each contact is on.", "Medium", "High", "High", "Campaigns"),

    # ═══════════════════════════════════ CAPABILITY ENGINE ═══════════════════════════════════
    ("C-01", "Capability Engine", "True Semantic Search with Embeddings", "Replace Jaccard similarity (which is just token overlap) with real transformer-based embeddings. Use OpenAI embeddings, Cohere, or a local model. Store vectors in a vector DB (pgvector, Pinecone, Qdrant).", "Critical", "High", "High", "Search"),
    ("C-02", "Capability Engine", "Vector Database Integration", "Add a vector store for capability assets. Generate embeddings on asset create/update. Perform approximate nearest neighbor (ANN) search. This is essential for RAG quality at scale.", "Critical", "High", "High", "Search"),
    ("C-03", "Capability Engine", "Add targetCompanySizes to Schema", "The search code reads cap.targetCompanySizes but the Prisma schema doesn't have this column. Add it as a String field (JSON array stored as string) so company size filtering actually works.", "Critical", "Medium", "Low", "Schema"),
    ("C-04", "Capability Engine", "Robust PDF Parsing", "Replace regex-based PDF text extraction (Tj/TJ operators) with a proper library like pdf-parse or pdfjs-dist. The current approach fails on most modern PDFs (compressed streams, encrypted).", "High", "High", "Low", "Ingestion"),
    ("C-05", "Capability Engine", "Robust DOCX Parsing", "Replace regex-based <w:t> tag parsing with the mammoth.js or docx4js library. Current approach fails on complex formatting, tables, headers/footers, and nested elements.", "High", "Medium", "Low", "Ingestion"),
    ("C-06", "Capability Engine", "Document Deduplication", "Hash extracted text (or use AI similarity) to detect duplicate uploads. Show warning before creating duplicate assets. Option to merge or replace existing assets.", "High", "Medium", "Medium", "Ingestion"),
    ("C-07", "Capability Engine", "Asset Versioning System", "The schema has a version field but the PUT endpoint never increments it. Add version tracking: auto-increment on edit, store previous versions, allow viewing/restoring history.", "Medium", "Medium", "Low", "CRUD"),
    ("C-08", "Capability Engine", "Asset Relationships & Linking", "Create explicit relationships between assets (e.g., a case_study links to its parent service_line, proof_points link to their case_study). Add parentAssetId or a junction table.", "High", "Medium", "Medium", "Data Model"),
    ("C-09", "Capability Engine", "Search Result Feedback Loop", "Add thumbs up/down on search results. Track which results were used in email generation. Use this signal to boost/reduce asset relevance scores over time.", "Medium", "High", "Medium", "Search"),
    ("C-10", "Capability Engine", "Knowledge Base Export/Import", "Add export to JSON/CSV and import from structured data. Users should be able to backup, share, and migrate their knowledge base between instances.", "Medium", "Medium", "Low", "Data Mgmt"),
    ("C-11", "Capability Engine", "Bulk Asset Operations", "Add bulk create, edit, delete, activate/deactivate for capability assets. Support CSV upload of structured knowledge data (not just documents).", "High", "Medium", "Low", "UX"),
    ("C-12", "Capability Engine", "Auto-Enrichment from Website", "Scrape the company's own website (services pages, case studies, about page) and auto-generate capability assets. Use LLM to extract structured knowledge from web content.", "Medium", "High", "High", "Ingestion"),
    ("C-13", "Capability Engine", "Knowledge Coverage Scoring Dashboard", "Enhance the coverage analysis to show a real-time score per industry, role, and service line. Visualize gaps as a heat map. Trigger alerts when coverage drops below thresholds.", "Medium", "Medium", "Medium", "Analytics"),
    ("C-14", "Capability Engine", "Multi-File Upload", "Allow uploading multiple documents at once (drag 5 PDFs). Process them in parallel. Show progress per file. Aggregate all extracted assets into the knowledge base.", "Medium", "Medium", "Medium", "UX"),
    ("C-15", "Capability Engine", "Asset Tagging System", "Add free-form tags to assets beyond the structured fields. Allow filtering and searching by tags. Support tag auto-suggestion based on existing tags.", "Low", "Low", "Low", "UX"),

    # ═══════════════════════════════════ LEAD ENGINE ═══════════════════════════════════
    ("L-01", "Lead Engine", "Unified Data Source (DB + Static)", "CRITICAL BUG FIXED: /api/leads was reading only from static JSON files while uploaded leads went to the DB. Now unified to auto-detect and serve from DB when available. Enhance with seamless merging.", "Critical", "High", "Done", "Architecture"),
    ("L-02", "Lead Engine", "Advanced Lead Scoring Model", "Current scoring is basic (role + email health + data completeness). Add: company fit score (industry match, size match), engagement score (opens, clicks, replies), recency decay, and AI-predicted conversion probability.", "High", "High", "Medium", "Scoring"),
    ("L-03", "Lead Engine", "Company Data Enrichment", "Integrate enrichment APIs (Clearbit, Apollo, Hunter.io, LinkedIn) to auto-populate: company size, revenue, tech stack, funding, employee count, social profiles. Store in CompanyResearchCard.", "High", "High", "Medium", "Enrichment"),
    ("L-04", "Lead Engine", "Lead Segmentation & Smart Lists", "Create dynamic segments based on criteria (industry, score range, status, location, engagement). Save segments for reuse. Show segment counts. Use segments as targeting for email campaigns.", "High", "High", "Medium", "Segmentation"),
    ("L-05", "Lead Engine", "Lead Export to CSV/Excel", "Add export button that downloads filtered leads as CSV or XLSX with all available fields. Include export options for selected leads only or entire filtered set.", "High", "Medium", "Low", "Data Mgmt"),
    ("L-06", "Lead Engine", "Lead Deduplication & Merge UI", "Add a duplicate detection view that shows potential duplicates (fuzzy name + email matching). Allow manual merge: pick primary, keep best fields, archive duplicates.", "High", "Medium", "Medium", "Data Quality"),
    ("L-07", "Lead Engine", "Status Workflow Management", "Implement the 10-state FSM (imported → cleaned → drafted → queued → sent → replied → bounced → suppressed → archived) with proper transitions, validation, and UI status badges.", "High", "Medium", "Medium", "Workflow"),
    ("L-08", "Lead Engine", "Column Mapping Preview & Edit", "Before import, show a preview table: detected column mappings, first 5 rows of data, let user edit/confirm mappings. Currently mapping is auto-detected with no user review.", "High", "Medium", "Medium", "Import"),
    ("L-09", "Lead Engine", "Incremental Import & Delta Processing", "For large files (>10k rows), process in background batches. Show real-time progress. Support resume on failure. Add import queue with priority.", "Medium", "Medium", "High", "Import"),
    ("L-10", "Lead Engine", "Real-time Email Verification Queue", "Currently verification is on-demand (single/bulk API). Add auto-verify on import with a background queue. Show verification progress. Flag unverifiable leads automatically.", "Medium", "Medium", "Medium", "Verification"),
    ("L-11", "Lead Engine", "Lead Activity Timeline", "Show a chronological timeline per lead: import date, verification result, drafts generated, emails sent, replies received, bounces, status changes. Single view of all activity.", "Medium", "Medium", "Medium", "UX"),
    ("L-12", "Lead Engine", "GDPR Consent Tracking", "Add proper consent fields: consent source (form, list purchase, manual), consent date, IP address, double opt-in support. Add consent status filter. Auto-suppress non-consented leads.", "High", "Medium", "Medium", "Compliance"),
    ("L-13", "Lead Engine", "Lead Notes & Activity Logging", "Add free-text notes per contact (CompanyNote exists but no ContactNote). Auto-log all actions (draft generated, email sent, status changed) to the AuditLog with full context.", "Medium", "Medium", "Low", "UX"),
    ("L-14", "Lead Engine", "Smart Lead Assignment", "For teams: assign leads to sales reps based on territory, industry expertise, or round-robin. Track ownership. Show rep performance per assigned lead.", "Medium", "Medium", "High", "Team Mgmt"),
    ("L-15", "Lead Engine", "Lead Source Tracking", "Track where each lead came from (LinkedIn, event, referral, cold list, inbound). Add source field to Contact. Analyze conversion rates by source in analytics.", "Medium", "Medium", "Low", "Analytics"),
]

# ── Create workbook ──
wb = Workbook()

# ═══════════════════════ SHEET 1: All Enhancements ═══════════════════════
ws = wb.active
ws.title = "Enhancement Roadmap"
ws.sheet_properties.tabColor = PRIMARY

headers = ["#", "ID", "Engine", "Enhancement", "Description", "Priority", "Impact", "Effort", "Category"]
col_widths = [4, 7, 16, 32, 70, 10, 10, 10, 14]
last_col = len(headers)

setup_sheet(ws, title="DeepMindQ — 45 Core Engine Enhancements", last_col=last_col)

# Header row (row 4)
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=i+1, value=h)
style_header_row(ws, 4, 2, last_col + 1)

# Column widths
ws.column_dimensions["A"].width = 3
for i, w in enumerate(col_widths):
    ws.column_dimensions[get_column_letter(i + 2)].width = w

# Priority fills
priority_fills = {
    "Critical": PatternFill("solid", fgColor="FDEDEC"),
    "High": PatternFill("solid", fgColor="FEF9E7"),
    "Medium": PatternFill("solid", fgColor="EBF5FB"),
    "Low": PatternFill("solid", fgColor="EAFAF1"),
}
priority_fonts = {
    "Critical": Font(name=font_body().name, size=11, color=ACCENT_NEGATIVE, bold=True),
    "High": Font(name=font_body().name, size=11, color=ACCENT_WARNING, bold=True),
    "Medium": Font(name=font_body().name, size=11, color="2980B9"),
    "Low": Font(name=font_body().name, size=11, color=ACCENT_POSITIVE),
}

# Engine separator fills
engine_colors = {
    "Email Engine": PatternFill("solid", fgColor="F4ECF7"),
    "Capability Engine": PatternFill("solid", fgColor="E8F8F5"),
    "Lead Engine": PatternFill("solid", fgColor="FEF5E7"),
}

row = 5
prev_engine = ""
for idx, enh in enumerate(enhancements):
    eid, engine, name, desc, priority, impact, effort, category = enh
    
    # Engine separator row
    if engine != prev_engine:
        if prev_engine:
            row += 1  # blank row between engines
        sep_fill = engine_colors.get(engine, fill_data_row(0))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col + 1)
        cell = ws.cell(row=row, column=2, value=f"  {engine}")
        cell.fill = sep_fill
        cell.font = Font(name=font_body().name, size=12, bold=True, color=PRIMARY)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28
        row += 1
        prev_engine = engine

    ws.cell(row=row, column=2, value=idx + 1).alignment = align_number()
    ws.cell(row=row, column=3, value=eid).alignment = align_text()
    ws.cell(row=row, column=4, value=engine).alignment = align_text()
    ws.cell(row=row, column=5, value=name).alignment = align_text()
    ws.cell(row=row, column=6, value=desc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=row, column=7, value=priority).alignment = align_header()
    ws.cell(row=row, column=8, value=impact).alignment = align_header()
    ws.cell(row=row, column=9, value=effort).alignment = align_header()
    ws.cell(row=row, column=10, value=category).alignment = align_text()

    # Apply row styling
    for col in range(2, last_col + 2):
        cell = ws.cell(row=row, column=col)
        cell.font = font_body()
        if col == 7:  # Priority column
            cell.fill = priority_fills.get(priority, fill_data_row(idx))
            cell.font = priority_fonts.get(priority, font_body())
        else:
            cell.fill = fill_data_row(idx)

    ws.row_dimensions[row].height = 52
    row += 1

# ═══════════════════════ SHEET 2: Priority Matrix ═══════════════════════
ws2 = wb.create_sheet("Priority Matrix")
ws2.sheet_properties.tabColor = ACCENT_NEGATIVE

# Count by priority per engine
matrix_data = {}
for eid, engine, name, desc, priority, impact, effort, category in enhancements:
    if engine not in matrix_data:
        matrix_data[engine] = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Total": 0}
    matrix_data[engine][priority] += 1
    matrix_data[engine]["Total"] += 1

setup_sheet(ws2, title="Priority Matrix — Engine Breakdown", last_col=6)

matrix_headers = ["Engine", "Critical", "High", "Medium", "Low", "Total"]
for i, h in enumerate(matrix_headers, 2):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 2, 7)

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 20
for c in ["C", "D", "E", "F", "G"]:
    ws2.column_dimensions[c].width = 12

r = 5
for engine, counts in matrix_data.items():
    ws2.cell(row=r, column=2, value=engine).font = font_body()
    ws2.cell(row=r, column=2).fill = fill_data_row(r - 5)
    for ci, key in enumerate(["Critical", "High", "Medium", "Low", "Total"], 3):
        cell = ws2.cell(row=r, column=ci, value=counts[key])
        cell.font = font_body()
        cell.alignment = align_number()
        cell.fill = fill_data_row(r - 5)
        if key == "Critical" and counts[key] > 0:
            cell.font = Font(name=font_body().name, size=11, color=ACCENT_NEGATIVE, bold=True)
        elif key == "Total":
            cell.font = Font(name=font_body().name, size=11, bold=True, color=PRIMARY)
    ws2.row_dimensions[r].height = 24
    r += 1

# Total row
ws2.cell(row=r, column=2, value="TOTAL").font = font_subheader()
for ci, key in enumerate(["Critical", "High", "Medium", "Low", "Total"], 3):
    total = sum(matrix_data[e][key] for e in matrix_data)
    cell = ws2.cell(row=r, column=ci, value=total)
    cell.font = font_subheader()
    cell.alignment = align_number()
    cell.fill = fill_header()
    cell.font = Font(name=font_body().name, size=11, bold=True, color="FFFFFF")
ws2.cell(row=r, column=2).fill = fill_header()
ws2.cell(row=r, column=2).font = Font(name=font_body().name, size=11, bold=True, color="FFFFFF")
ws2.row_dimensions[r].height = 28

# ═══════════════════════ SHEET 3: Quick Wins ═══════════════════════
ws3 = wb.create_sheet("Quick Wins")
ws3.sheet_properties.tabColor = ACCENT_POSITIVE

# Filter: High Impact + Low/Medium Effort
quick_wins = [e for e in enhancements if e[5] == "High" and e[6] in ("Low", "Medium")]

setup_sheet(ws3, title="Quick Wins — High Impact, Lower Effort", last_col=6)

qw_headers = ["#", "ID", "Enhancement", "Engine", "Impact", "Effort"]
for i, h in enumerate(qw_headers, 2):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 2, 7)

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 4
ws3.column_dimensions["C"].width = 7
ws3.column_dimensions["D"].width = 38
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 10
ws3.column_dimensions["G"].width = 10

r = 5
for idx, (eid, engine, name, desc, priority, impact, effort, category) in enumerate(quick_wins):
    ws3.cell(row=r, column=2, value=idx + 1).alignment = align_number()
    ws3.cell(row=r, column=3, value=eid).alignment = align_text()
    ws3.cell(row=r, column=4, value=name).alignment = align_text()
    ws3.cell(row=r, column=5, value=engine).alignment = align_text()
    ws3.cell(row=r, column=6, value=impact).alignment = align_header()
    ws3.cell(row=r, column=7, value=effort).alignment = align_header()
    
    for col in range(2, 8):
        cell = ws3.cell(row=r, column=col)
        cell.font = font_body()
        cell.fill = fill_data_row(idx)
    ws3.row_dimensions[r].height = 24
    r += 1

# ═══════════════════════ SHEET 4: Implementation Phases ═══════════════════════
ws4 = wb.create_sheet("Implementation Phases")
ws4.sheet_properties.tabColor = "2980B9"

phases = [
    ("Phase 1: Critical Fixes (Week 1-2)", [
        ("E-01", "SMTP Provider Integration", "Email Engine"),
        ("E-02", "Queue PATCH Endpoint", "Email Engine"),
        ("E-07", "Inbound Email Webhook", "Email Engine"),
        ("E-09", "Unsubscribe Link & Compliance", "Email Engine"),
        ("C-01", "True Semantic Search with Embeddings", "Capability Engine"),
        ("C-03", "Add targetCompanySizes to Schema", "Capability Engine"),
        ("L-01", "Unified Data Source (Already Fixed)", "Lead Engine"),
    ]),
    ("Phase 2: Core Features (Week 3-5)", [
        ("E-03", "Email Sending Cron Worker", "Email Engine"),
        ("E-04", "Retry Logic with Backoff", "Email Engine"),
        ("E-05", "Email Scheduling System", "Email Engine"),
        ("E-08", "Bounce Webhook Handler", "Email Engine"),
        ("E-12", "Bulk Draft Operations", "Email Engine"),
        ("E-13", "Deduplicate Generation Code", "Email Engine"),
        ("E-14", "Email Template System", "Email Engine"),
        ("C-04", "Robust PDF Parsing", "Capability Engine"),
        ("C-05", "Robust DOCX Parsing", "Capability Engine"),
        ("C-08", "Asset Relationships & Linking", "Capability Engine"),
        ("C-11", "Bulk Asset Operations", "Capability Engine"),
        ("L-02", "Advanced Lead Scoring Model", "Lead Engine"),
        ("L-03", "Company Data Enrichment", "Lead Engine"),
        ("L-04", "Lead Segmentation & Smart Lists", "Lead Engine"),
        ("L-05", "Lead Export to CSV/Excel", "Lead Engine"),
        ("L-07", "Status Workflow Management", "Lead Engine"),
        ("L-08", "Column Mapping Preview & Edit", "Lead Engine"),
    ]),
    ("Phase 3: Optimization (Week 6-8)", [
        ("E-06", "Email Thread Tracking", "Email Engine"),
        ("E-10", "A/B Testing for Subject Lines", "Email Engine"),
        ("E-11", "Email Open & Click Tracking", "Email Engine"),
        ("E-15", "Multi-Step Email Sequences", "Email Engine"),
        ("C-02", "Vector Database Integration", "Capability Engine"),
        ("C-06", "Document Deduplication", "Capability Engine"),
        ("C-07", "Asset Versioning System", "Capability Engine"),
        ("C-09", "Search Result Feedback Loop", "Capability Engine"),
        ("C-10", "Knowledge Base Export/Import", "Capability Engine"),
        ("C-12", "Auto-Enrichment from Website", "Capability Engine"),
        ("C-13", "Knowledge Coverage Scoring Dashboard", "Capability Engine"),
        ("C-14", "Multi-File Upload", "Capability Engine"),
        ("L-06", "Lead Deduplication & Merge UI", "Lead Engine"),
        ("L-09", "Incremental Import & Delta Processing", "Lead Engine"),
        ("L-10", "Real-time Email Verification Queue", "Lead Engine"),
        ("L-11", "Lead Activity Timeline", "Lead Engine"),
        ("L-12", "GDPR Consent Tracking", "Lead Engine"),
    ]),
    ("Phase 4: Scale & Polish (Week 9+)", [
        ("C-15", "Asset Tagging System", "Capability Engine"),
        ("L-13", "Lead Notes & Activity Logging", "Lead Engine"),
        ("L-14", "Smart Lead Assignment", "Lead Engine"),
        ("L-15", "Lead Source Tracking", "Lead Engine"),
    ]),
]

setup_sheet(ws4, title="Suggested Implementation Phases", last_col=4)

phase_headers = ["Phase", "ID", "Enhancement", "Engine"]
for i, h in enumerate(phase_headers, 2):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 2, 5)

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 36
ws4.column_dimensions["C"].width = 7
ws4.column_dimensions["D"].width = 38
ws4.column_dimensions["E"].width = 18

r = 5
phase_idx = 0
for phase_name, items in phases:
    # Phase header
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    cell = ws4.cell(row=r, column=2, value=phase_name)
    cell.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    cell.font = Font(name=font_body().name, size=11, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[r].height = 28
    r += 1
    
    for eid, ename, eengine in items:
        ws4.cell(row=r, column=3, value=eid).alignment = align_text()
        ws4.cell(row=r, column=4, value=ename).alignment = align_text()
        ws4.cell(row=r, column=5, value=eengine).alignment = align_text()
        for col in range(2, 6):
            cell = ws4.cell(row=r, column=col)
            cell.font = font_body()
            cell.fill = fill_data_row(r - 5)
        ws4.row_dimensions[r].height = 22
        r += 1
    r += 1  # gap between phases
    phase_idx += 1

# ── Save ──
output_path = "/home/z/my-project/download/DeepMindQ_45_Enhancement_Roadmap.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Saved to {output_path}")