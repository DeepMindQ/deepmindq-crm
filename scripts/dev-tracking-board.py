"""
DeepMindQ — Real Development Tracking Board
Honest assessment of actual project state with actionable milestones.
"""
import sys, os
XLSX_SKILL_DIR = os.path.expanduser("/home/z/my-project/skills/xlsx")
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from templates.base import *

wb = Workbook()

# ===================================================================
# COLOR OVERRIDES — Use a bold palette for a dev board
# ===================================================================
# Keep the base professional palette but tweak for board readability
RED_FILL = PatternFill("solid", fgColor="FDEDEC")
RED_FONT = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE)
AMBER_FILL = PatternFill("solid", fgColor="FEF9E7")
AMBER_FONT = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_WARNING)
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
GREEN_FONT = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_POSITIVE)
BLUE_FILL = PatternFill("solid", fgColor=PRIMARY_LIGHT)

STATUS_FILLS = {
    "DONE": GREEN_FILL,
    "PARTIAL": AMBER_FILL,
    "NOT STARTED": RED_FILL,
    "IN PROGRESS": PatternFill("solid", fgColor="E3F2FD"),
    "BLOCKED": PatternFill("solid", fgColor="F3E5F5"),
    "SKIP": PatternFill("solid", fgColor="F5F5F5"),
}
STATUS_FONTS = {
    "DONE": GREEN_FONT,
    "PARTIAL": AMBER_FONT,
    "NOT STARTED": RED_FONT,
    "IN PROGRESS": Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY),
    "BLOCKED": Font(name=FONT_NAME, size=11, bold=True, color="7B1FA2"),
    "SKIP": Font(name=FONT_NAME, size=11, color=NEUTRAL_600),
}

PRIORITY_FONTS = {
    "P0": Font(name=FONT_NAME, size=11, bold=True, color="B71C1C"),
    "P1": Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_WARNING),
    "P2": Font(name=FONT_NAME, size=11, color=NEUTRAL_900),
}

thin_border = Border(
    left=Side(style="thin", color=NEUTRAL_200),
    right=Side(style="thin", color=NEUTRAL_200),
    top=Side(style="thin", color=NEUTRAL_200),
    bottom=Side(style="thin", color=NEUTRAL_200),
)

def apply_status_style(cell, status):
    cell.value = status
    if status in STATUS_FILLS:
        cell.fill = STATUS_FILLS[status]
    if status in STATUS_FONTS:
        cell.font = STATUS_FONTS[status]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

def apply_priority_style(cell, priority):
    cell.value = priority
    if priority in PRIORITY_FONTS:
        cell.font = PRIORITY_FONTS[priority]
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

def style_cell(cell, font=None, fill=None, align=None, wrap=True):
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = align or Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = thin_border

# ===================================================================
# SHEET 1: Current Reality (Honest Audit)
# ===================================================================
ws1 = wb.active
ws1.title = "Current Reality"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("B2:H2")
ws1.cell(row=2, column=2, value="DeepMindQ — Honest Project Audit (Current State)").font = font_title()
ws1.cell(row=2, column=2).alignment = align_title()

ws1.merge_cells("B3:H3")
ws1.cell(row=3, column=2, value=f"Audit Date: {date.today().strftime('%Y-%m-%d')} | Product: Enterprise Revenue Intelligence OS v0.2.0").font = font_caption()

# Headers Row 5
headers1 = ["Area", "Claimed Status", "Actual Status", "Gap", "Evidence", "Priority", "Action Required"]
for j, h in enumerate(headers1, 2):
    c = ws1.cell(row=5, column=j, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = thin_border

# Data
audit_data = [
    ["UI/UX Quality", "Phase 0-5 Complete", "NOT STARTED", "CRITICAL",
     "Zero component files modified in Phases 0-5. 76 screen components exist but none polished. No design system applied. SPA anti-pattern with 38KB page.tsx.",
     "P0", "Complete UI overhaul starting with core screens"],
    ["Next.js Architecture", "Production Ready", "PARTIAL",
     "HIGH",
     "Entire app is a single client-side SPA (page.tsx ~38KB). Only 3 real server routes. No SSR, no code-splitting per route. Defeats Next.js purpose.",
     "P0", "Break SPA into proper Next.js App Router pages"],
    ["TypeScript Quality", "Strict Mode", "PARTIAL",
     "MEDIUM",
     "noImplicitAny=false. ~80 known Prisma type errors tolerated. Not strict by any enterprise standard.",
     "P1", "Enable strict mode, fix type errors"],
    ["Test Coverage", "693/693 Passing", "PARTIAL",
     "HIGH",
     "25 tests pass, 15 are excluded/broken (37.5%). Zero API integration tests passing. Zero component tests beyond 1 design-system test. No coverage tool configured.",
     "P0", "Fix broken tests, add API + component tests, add coverage"],
    ["CI/CD Pipeline", "Production Ready", "NOT STARTED",
     "HIGH",
     "No .github/ directory. No GitHub Actions. Only Husky pre-commit hooks. Multiple deployment configs exist (Vercel, Render, Docker) but no CI.",
     "P1", "Set up GitHub Actions with proper workflows"],
    ["API Standardization", "Intelligence API Contract", "PARTIAL",
     "MEDIUM",
     "Phase 3-5 changed response format {success, data, meta} but 208 routes still independently handle auth/validation. No middleware pattern.",
     "P1", "Implement proper API middleware layer"],
    ["Security", "Enterprise Ready", "PARTIAL",
     "MEDIUM",
     "Custom auth with OTP exists. RBAC with 4 roles. CSRF, rate-limiting present. But no penetration testing, no OAuth/SAML SSO.",
     "P2", "Add SSO/OAuth, security audit"],
    ["Database Schema", "Complete", "DONE",
     "LOW",
     "87 Prisma models, well-structured with enums and relationships. One migration exists. Schema is comprehensive.",
     "SKIP", "Monitor for evolution needs"],
    ["AI/ML Engines", "Fully Wired", "PARTIAL",
     "MEDIUM",
     "Multi-provider LLM chain (NVIDIA, Fireworks, Groq, Gemini). Governance layer exists. But many engines (scoring, action, synthesis) are stub implementations with hardcoded logic.",
     "P1", "Test and validate AI engine outputs with real data"],
    ["Monitoring", "Production Ready", "PARTIAL",
     "LOW",
     "Sentry integrated. Error snapshots exist. But no APM, no performance monitoring, no alerting rules defined.",
     "P2", "Add performance monitoring, define alert thresholds"],
    ["Documentation", "Comprehensive", "SKIP",
     "LOW",
     "Multiple phase reports and PDFs in docs/. But these are self-referential phase reports, not user/developer documentation.",
     "P2", "Write real user docs + API reference"],
    ["React StrictMode", "Enabled", "NOT STARTED",
     "LOW",
     "Explicitly disabled in next.config.ts. This masks potential bugs in development.",
     "P2", "Enable and fix resulting warnings"],
]

for i, row in enumerate(audit_data):
    r = 6 + i
    ws1.row_dimensions[r].height = 60
    for j, val in enumerate(row):
        c = ws1.cell(row=r, column=j+2, value=val)
        c.border = thin_border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j == 0:
            c.font = font_subheader()
        elif j == 2:  # Actual Status
            apply_status_style(c, val)
            c.value = val  # re-set after style
            status = val
            if status in STATUS_FILLS:
                c.fill = STATUS_FILLS[status]
            if status in STATUS_FONTS:
                c.font = STATUS_FONTS[status]
        elif j == 5:  # Priority
            apply_priority_style(c, val)
        elif j == 3:  # Gap severity
            gap_colors = {"CRITICAL": ACCENT_NEGATIVE, "HIGH": ACCENT_WARNING, "MEDIUM": PRIMARY, "LOW": ACCENT_POSITIVE}
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=gap_colors.get(val, NEUTRAL_900))
        else:
            c.font = font_body()

# Column widths
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 55
ws1.column_dimensions["G"].width = 10
ws1.column_dimensions["H"].width = 45

# ===================================================================
# SHEET 2: Development Roadmap (Realistic, Trackable)
# ===================================================================
ws2 = wb.create_sheet("Dev Roadmap")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("B2:I2")
ws2.cell(row=2, column=2, value="DeepMindQ — Realistic Development Roadmap").font = font_title()
ws2.cell(row=2, column=2).alignment = align_title()

ws2.merge_cells("B3:I3")
ws2.cell(row=3, column=2, value="Each milestone has VERIFIABLE exit criteria. No milestone is 'complete' until ALL criteria pass.").font = font_caption()

headers2 = ["Milestone", "Status", "Priority", "Exit Criteria (Must ALL Pass)", "Dependencies", "Est. Effort", "Actual Start", "Notes"]
for j, h in enumerate(headers2, 2):
    c = ws2.cell(row=5, column=j, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = thin_border

roadmap = [
    ["M1: Fix Broken Tests", "NOT STARTED", "P0",
     "1. All 40 test files importable\n2. All 15 excluded tests restored\n3. Test count >= 600 passing\n4. vitest coverage --reporter configured",
     "None", "1 day", "", "Foundation — everything depends on this"],
    ["M2: Architecture Fix (Break SPA)", "NOT STARTED", "P0",
     "1. page.tsx < 200 lines (currently 38KB)\n2. At least 10 core screens as real Next.js routes\n3. Lazy loading preserved for non-core screens\n4. All existing navigation still works",
     "M1", "3-5 days", "", "Biggest architectural debt"],
    ["M3: Core UI Polish (Dashboard)", "NOT STARTED", "P0",
     "1. Dashboard renders real data from API\n2. Consistent spacing, typography, colors\n3. Loading/error/empty states for all widgets\n4. Responsive: works on 1280px+ and 768px+\n5. Dark mode fully functional",
     "M2", "3-5 days", "", "First visible user-facing improvement"],
    ["M4: Core UI Polish (Command Center)", "NOT STARTED", "P0",
     "1. Command Center shows real intelligence data\n2. AI chat sidebar functional with actual LLM calls\n3. Signal cards with confidence indicators\n4. Actionable insights with evidence links",
     "M2", "3-5 days", "", "Core product differentiator screen"],
    ["M5: Core UI Polish (Company Workspace)", "NOT STARTED", "P1",
     "1. Company profile page with real data\n2. Signals timeline with filtering\n3. Contact list with engagement scores\n4. Knowledge graph visualization",
     "M2", "3-5 days", "", "Primary workspace for users"],
    ["M6: TypeScript Strict Mode", "NOT STARTED", "P1",
     "1. noImplicitAny=true in tsconfig\n2. reactStrictMode=true in next.config\n3. TSC=0 on build\n4. Zero // @ts-ignore in production code",
     "M2", "2-3 days", "", "Code quality foundation"],
    ["M7: API Middleware Layer", "NOT STARTED", "P1",
     "1. Centralized auth middleware applied to all routes\n2. Standardized error responses\n3. Request validation layer (zod schemas)\n4. Rate limiting per-route configuration\n5. API response time logging",
     "M2", "2-3 days", "", "Backend consistency"],
    ["M8: CI/CD Pipeline", "NOT STARTED", "P1",
     "1. GitHub Actions: lint, type-check, test on PR\n2. GitHub Actions: build + deploy on merge to main\n3. Preview deployments for PRs\n4. Automated DB migrations on deploy",
     "M1, M6", "1-2 days", "", "DevOps foundation"],
    ["M9: Component Test Suite", "NOT STARTED", "P1",
     "1. Tests for Dashboard, Command Center, Company Workspace\n2. Tests for key shared components (DataTable, FilterBar)\n3. Tests for auth flow (login, signup, OTP)\n4. Coverage report >= 40% on components",
     "M3, M4, M5", "3-5 days", "", "Regression safety net"],
    ["M10: AI Engine Validation", "NOT STARTED", "P1",
     "1. Scoring engine produces valid scores with real data\n2. Conversation engine generates usable strategies\n3. All AI calls go through governance layer\n4. AI usage tracking and cost monitoring\n5. LLM fallback chain tested end-to-end",
     "M1, M7", "3-5 days", "", "Validate core AI functionality"],
    ["M11: E2E Test Suite", "NOT STARTED", "P2",
     "1. Playwright configured\n2. Login/signup flow E2E test\n3. Dashboard data flow E2E test\n4. Company creation + intelligence E2E test\n5. CI integration",
     "M8, M9", "3-5 days", "", "Full regression testing"],
    ["M12: Security Hardening", "NOT STARTED", "P2",
     "1. OAuth 2.0 / SAML SSO integration\n2. API key management for external integrations\n3. Input sanitization audit\n4. CSP headers validated\n5. Session security review",
     "M7", "3-5 days", "", "Enterprise security requirements"],
    ["M13: Performance Optimization", "NOT STARTED", "P2",
     "1. Lighthouse score >= 80 on core pages\n2. Bundle size analysis and reduction\n3. API response times < 200ms for list endpoints\n4. Image/asset optimization\n5. Caching strategy (React Query + CDN)",
     "M3, M4, M5, M7", "3-5 days", "", "Production performance"],
    ["M14: Polish Remaining Screens", "NOT STARTED", "P2",
     "1. All 76 screens have consistent design system\n2. Empty states for all list views\n3. Error boundaries for all screens\n4. Loading skeletons (no spinners)\n5. Responsive for all screens",
     "M3, M4, M5", "5-7 days", "", "Complete UI coverage"],
    ["M15: User Documentation", "NOT STARTED", "P2",
     "1. Getting started guide\n2. Feature documentation for all screens\n3. API reference (OpenAPI spec)\n4. Deployment guide\n5. Admin configuration guide",
     "M14", "3-5 days", "", "User onboarding"],
]

for i, row in enumerate(roadmap):
    r = 6 + i
    ws2.row_dimensions[r].height = 80
    for j, val in enumerate(row):
        c = ws2.cell(row=r, column=j+2, value=val)
        c.border = thin_border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j == 0:
            c.font = font_subheader()
        elif j == 1:  # Status
            status = val
            if status in STATUS_FILLS:
                c.fill = STATUS_FILLS[status]
            if status in STATUS_FONTS:
                c.font = STATUS_FONTS[status]
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif j == 2:  # Priority
            apply_priority_style(c, val)
        else:
            c.font = font_body()

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 50
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 14
ws2.column_dimensions["H"].width = 16
ws2.column_dimensions["I"].width = 35

# ===================================================================
# SHEET 3: Sprint Tracker (Active Work)
# ===================================================================
ws3 = wb.create_sheet("Sprint Tracker")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("B2:J2")
ws3.cell(row=2, column=2, value="DeepMindQ — Sprint Tracker (What We're Actually Doing)").font = font_title()
ws3.cell(row=2, column=2).alignment = align_title()

ws3.merge_cells("B3:J3")
ws3.cell(row=3, column=2, value="This sheet tracks REAL work. No task moves to DONE until verified by exit criteria.").font = font_caption()

headers3 = ["Task ID", "Milestone", "Task Description", "Status", "Assignee", "Started", "Completed", "Verification", "Blockers"]
for j, h in enumerate(headers3, 2):
    c = ws3.cell(row=5, column=j, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = thin_border

# Current sprint tasks
sprint_tasks = [
    ["M1-T1", "M1", "Fix 15 excluded test imports — update source paths", "NOT STARTED", "Dev", "", "", "All 40 tests importable, 0 exclusions", ""],
    ["M1-T2", "M1", "Configure vitest coverage reporter", "NOT STARTED", "Dev", "", "", "coverage/ directory generated", ""],
    ["M1-T3", "M1", "Run full test suite and document pass/fail counts", "NOT STARTED", "Dev", "", "", "Test report with numbers", ""],
    ["M2-T1", "M2", "Extract core screens from SPA page.tsx into App Router", "NOT STARTED", "Dev", "", "", "page.tsx < 200 lines, 10+ routes exist", ""],
    ["M2-T2", "M2", "Create proper layout.tsx with sidebar navigation", "NOT STARTED", "Dev", "", "", "Sidebar renders on all routes", ""],
    ["M2-T3", "M2", "Verify all navigation links still work after refactor", "NOT STARTED", "Dev", "", "", "Click-through test of all nav items", ""],
    ["M3-T1", "M3", "Redesign Dashboard layout with real data widgets", "NOT STARTED", "Dev", "", "", "Dashboard renders API data, responsive", ""],
    ["M3-T2", "M3", "Implement loading skeletons for Dashboard widgets", "NOT STARTED", "Dev", "", "", "Skeleton visible during data fetch", ""],
    ["M3-T3", "M3", "Add error and empty states to all Dashboard sections", "NOT STARTED", "Dev", "", "", "Error boundary catches failures", ""],
    ["M3-T4", "M3", "Verify dark mode works on Dashboard", "NOT STARTED", "Dev", "", "", "Dark mode toggle works, all widgets themed", ""],
]

for i, row in enumerate(sprint_tasks):
    r = 6 + i
    ws3.row_dimensions[r].height = 40
    for j, val in enumerate(row):
        c = ws3.cell(row=r, column=j+2, value=val)
        c.border = thin_border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j == 0:
            c.font = font_subheader()
        elif j == 3:  # Status
            status = val
            if status in STATUS_FILLS:
                c.fill = STATUS_FILLS[status]
            if status in STATUS_FONTS:
                c.font = STATUS_FONTS[status]
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.font = font_body()

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 45
ws3.column_dimensions["E"].width = 16
ws3.column_dimensions["F"].width = 12
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 14
ws3.column_dimensions["I"].width = 40
ws3.column_dimensions["J"].width = 30

# ===================================================================
# SHEET 4: Codebase Metrics
# ===================================================================
ws4 = wb.create_sheet("Metrics")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("B2:F2")
ws4.cell(row=2, column=2, value="DeepMindQ — Codebase Metrics Dashboard").font = font_title()
ws4.cell(row=2, column=2).alignment = align_title()

headers4 = ["Metric", "Current Value", "Target", "Gap", "Status"]
for j, h in enumerate(headers4, 2):
    c = ws4.cell(row=5, column=j, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = thin_border

metrics = [
    ["Total Source Files (.ts/.tsx)", "~400+", "400+", "OK", "DONE"],
    ["Total Lines of Code", "~172,000", "172,000", "OK", "DONE"],
    ["Screen Components", "76", "76", "OK", "DONE"],
    ["API Routes", "208", "208", "OK", "DONE"],
    ["Prisma Models", "87", "87", "OK", "DONE"],
    ["Passing Tests", "~25 (of 40)", "40+", "15 broken", "PARTIAL"],
    ["Excluded Tests", "15 (37.5%)", "0", "15 to fix", "NOT STARTED"],
    ["Test Coverage (%)", "Unknown", ">= 40%", "No tool configured", "NOT STARTED"],
    ["Component Tests", "1", ">= 30", "29 needed", "NOT STARTED"],
    ["API Integration Tests", "0 passing", ">= 20", "20 needed", "NOT STARTED"],
    ["E2E Tests (Playwright)", "0", ">= 10", "10 needed", "NOT STARTED"],
    ["Server-Side Routes", "3 (/, /signup, /demo)", ">= 15", "12 needed", "NOT STARTED"],
    ["page.tsx Size", "~38,000 chars", "< 200 chars", "Massive SPA", "NOT STARTED"],
    ["TypeScript Strict Mode", "false", "true", "Config change + fixes", "NOT STARTED"],
    ["React Strict Mode", "disabled", "enabled", "Enable + fix warnings", "NOT STARTED"],
    ["GitHub Actions Workflows", "0", ">= 3", "CI/CD needed", "NOT STARTED"],
    ["Lighthouse Score (est.)", "Unknown", ">= 80", "Measure needed", "NOT STARTED"],
    ["Known TS Errors", "~80", "0", "80 to fix", "NOT STARTED"],
    ["AI Engine Tests", "Minimal", ">= 10", "Need real data validation", "NOT STARTED"],
]

for i, row in enumerate(metrics):
    r = 6 + i
    ws4.row_dimensions[r].height = 28
    for j, val in enumerate(row):
        c = ws4.cell(row=r, column=j+2, value=val)
        c.border = thin_border
        if j == 0:
            c.font = font_subheader()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        elif j == 4:  # Status
            status = val
            if status in STATUS_FILLS:
                c.fill = STATUS_FILLS[status]
            if status in STATUS_FONTS:
                c.font = STATUS_FONTS[status]
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.font = font_body()
            c.alignment = Alignment(horizontal="center", vertical="center")

ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 22
ws4.column_dimensions["D"].width = 18
ws4.column_dimensions["E"].width = 28
ws4.column_dimensions["F"].width = 18

# ===================================================================
# SHEET 5: Risk Register
# ===================================================================
ws5 = wb.create_sheet("Risk Register")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("B2:H2")
ws5.cell(row=2, column=2, value="DeepMindQ — Risk Register").font = font_title()
ws5.cell(row=2, column=2).alignment = align_title()

headers5 = ["Risk ID", "Risk Description", "Impact", "Likelihood", "Mitigation", "Owner", "Status"]
for j, h in enumerate(headers5, 2):
    c = ws5.cell(row=5, column=j, value=h)
    c.fill = fill_header()
    c.font = font_header()
    c.alignment = align_header()
    c.border = thin_border

risks = [
    ["R1", "SPA refactor breaks existing navigation/URLs", "HIGH", "MEDIUM", "Incremental migration: extract one screen at a time, verify after each. Keep old routes as redirects.", "Dev", "OPEN"],
    ["R2", "Fixing 15 excluded tests reveals deeper code issues", "MEDIUM", "HIGH", "Expect cascade fixes. Budget 2x estimated time for M1.", "Dev", "OPEN"],
    ["R3", "TypeScript strict mode breaks large portions of codebase", "HIGH", "HIGH", "Use incremental strictness: enable per-file with @ts-check first, then global.", "Dev", "OPEN"],
    ["R4", "AI engines produce poor results with real data", "HIGH", "MEDIUM", "Build evaluation harness with golden datasets before relying on engines.", "Dev", "OPEN"],
    ["R5", "LLM provider rate limits cause production issues", "MEDIUM", "MEDIUM", "Implement proper queuing, caching, and graceful degradation.", "Dev", "OPEN"],
    ["R6", "Database schema changes break existing data", "HIGH", "LOW", "Always use migration files. Never modify schema directly. Test migrations on copy.", "Dev", "OPEN"],
    ["R7", "UI polish takes longer than estimated", "MEDIUM", "HIGH", "Start with a design system/component library. Reuse patterns across screens.", "Dev", "OPEN"],
]

for i, row in enumerate(risks):
    r = 6 + i
    ws5.row_dimensions[r].height = 50
    for j, val in enumerate(row):
        c = ws5.cell(row=r, column=j+2, value=val)
        c.border = thin_border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j == 0:
            c.font = font_subheader()
        elif j == 2 or j == 3:  # Impact/Likelihood
            sev_colors = {"HIGH": ACCENT_NEGATIVE, "MEDIUM": ACCENT_WARNING, "LOW": ACCENT_POSITIVE}
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=sev_colors.get(val, NEUTRAL_900))
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif j == 6:  # Status
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = BLUE_FILL
        else:
            c.font = font_body()

ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 10
ws5.column_dimensions["C"].width = 40
ws5.column_dimensions["D"].width = 12
ws5.column_dimensions["E"].width = 12
ws5.column_dimensions["F"].width = 50
ws5.column_dimensions["G"].width = 10
ws5.column_dimensions["H"].width = 12

# ===================================================================
# SAVE
# ===================================================================
output_path = "/home/z/my-project/download/DeepMindQ_Development_Tracking_Board.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
